from ch00_py.file_toolbox import count_dirs_files, create_path
from ch19_idea_src.idea2brick import (
    IdeaBook,
    SheetRef,
    create_spark_face_spark_nums,
    get_idea_config_dict,
    get_idea_sheet_refs,
    get_max_spark_num_from_files,
    get_spark_faces_from_df,
    get_spark_faces_from_files,
    ideas_sheets_to_brick_sheets,
    set_spark_num_column,
    validate_idea_columns,
)
from ch99_glossary.keywords import Ch19Keywords as kw, ExampleStrs as exx
from openpyxl import Workbook as openpyxl_Workbook
from os.path import join as os_path_join
from pandas import (
    DataFrame as pandas_DataFrame,
    ExcelWriter as pandas_ExcelWriter,
    isna as pandas_isna,
    read_excel as pandas_read_excel,
)
from pathlib import Path
from pytest import fixture as pytest_fixture, raises as pytest_raises


def test_IdeaBook_Exists():
    # ESTABLISH / WHEN
    ideabook = IdeaBook()
    # THEN
    assert not ideabook.ideas
    assert set(ideabook.__dict__.keys()) == {f"{kw.idea}s"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario0_Basic():
    # ESTABLISH
    df = pandas_DataFrame({kw.spark_face: ["a", "b", "a", "c"]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario1_excludes_nulls():
    # ESTABLISH
    df = pandas_DataFrame({kw.spark_face: ["a", None, "b", float("nan")]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == {"a", "b"}


def test_get_spark_faces_from_df_ReturnsObj_Scenario2_MissingColumnReturnsEmptySet():
    # ESTABLISH
    df = pandas_DataFrame({"other_col": [1, 2, 3]})
    # WHEN
    result = get_spark_faces_from_df(df)
    # THEN
    assert result == set()


def test_get_spark_faces_from_files_ReturnsObj_Scenario0_Multiple_files(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"

    df1 = pandas_DataFrame({kw.spark_face: ["a", "b"]})
    df2 = pandas_DataFrame({kw.spark_face: ["b", "c"]})

    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", "b", "c"}


def test_get_spark_faces_from_files_ReturnsObj_Scenario1_Multiple_sheets(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = pandas_DataFrame({kw.spark_face: ["a"]})
    df2 = pandas_DataFrame({kw.spark_face: [exx.sue]})

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a", exx.sue}


def test_get_spark_faces_from_files_ReturnsObj_Scenario2_IgnoresMissingColumn(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"

    df1 = pandas_DataFrame({kw.spark_face: ["a"]})
    df2 = pandas_DataFrame({"other": [1, 2]})  # no spark_face column

    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
    # WHEN
    result = get_spark_faces_from_files(tmp_path)
    # THEN
    assert result == {"a"}


def test_get_max_spark_num_from_files_ReturnsObj_Scenario0_MultipleFiles(tmp_path):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    file2 = tmp_path / "file2.xlsx"
    df1 = pandas_DataFrame({kw.spark_num: [1, 2, 3]})
    df2 = pandas_DataFrame({kw.spark_num: [4, 5]})
    df1.to_excel(file1, index=False)
    df2.to_excel(file2, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 5


def test_get_max_spark_num_from_files_ReturnsObj_Scenario1_IgnoresInvalidAndConvertsFloats(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df = pandas_DataFrame({kw.spark_num: ["10", "bad", None, 7.9]})  # 7.9 -> 7
    df.to_excel(file1, index=False)
    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 10


def test_get_max_spark_num_from_files_ReturnsObj_Scenario2_MultipleSheetsAndMissingColumn(
    tmp_path,
):
    # ESTABLISH
    file1 = tmp_path / "file1.xlsx"
    df1 = pandas_DataFrame({kw.spark_num: [1, 20]})
    df2 = pandas_DataFrame({"other": [100, 200]})  # no spark_num
    df3 = pandas_DataFrame({kw.spark_num: [15]})
    with pandas_ExcelWriter(file1) as writer:
        df1.to_excel(writer, sheet_name="Sheet1", index=False)
        df2.to_excel(writer, sheet_name="Sheet2", index=False)
        df3.to_excel(writer, sheet_name="Sheet3", index=False)

    # WHEN
    result = get_max_spark_num_from_files(tmp_path)
    # THEN
    assert result == 20


def test_create_spark_face_spark_nums_ReturnsObj_Scenario0_Simple():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = 11
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 12, exx.sue: 13, exx.yao: 14}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario1_max_spark_num_IsNone():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    max_spark_num = None
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces, max_spark_num)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_create_spark_face_spark_nums_ReturnsObj_Scenario2_No_max_spark_num():
    # ESTABLISH
    spark_faces = {exx.sue, exx.bob, exx.yao}
    # WHEN
    x_dict = create_spark_face_spark_nums(spark_faces)
    # THEN
    assert x_dict == {exx.bob: 1, exx.sue: 2, exx.yao: 3}


def test_set_spark_num_column_SetsAttr_Scenario0_Add_spark_num_Basic():
    # ESTABLISH
    df = pandas_DataFrame({kw.spark_face: ["a", "b", "c"]})
    mapping = {"a": 1, "b": 2, "c": 3}
    # WHEN
    set_spark_num_column(df, mapping)
    # THEN
    assert list(df.columns)[0] == kw.spark_num
    assert df[kw.spark_num].tolist() == [1, 2, 3]


def test_set_spark_num_column_SetsAttr_Scenario1_MissingSparkFaceSets_nan():
    # ESTABLISH
    df = pandas_DataFrame({kw.spark_face: ["a", "b", "x"]})
    mapping = {"a": 1, "b": 2}
    # WHEN
    set_spark_num_column(df, mapping)
    # THEN
    assert df[kw.spark_num].tolist()[:2] == [1, 2]
    assert pandas_isna(df[kw.spark_num].iloc[2])


def test_set_spark_num_column_SetsAttr_Scenario0_MutatesOriginalDataframe():
    # ESTABLISH
    df = pandas_DataFrame({kw.spark_face: ["a", "b"]})
    mapping = {"a": 1, "b": 2}
    assert kw.spark_num not in df.columns
    # WHEN
    set_spark_num_column(df, mapping)
    # THEN
    assert kw.spark_num in df.columns


def test_SheetRef_Exists():
    # ESTABLISH
    x_src_filename = "file_x.xlsx"
    bk_sheet_name = "bk00005"
    # WHEN
    sheet_ref = SheetRef(x_src_filename, bk_sheet_name)
    # THEN
    assert sheet_ref.src_filename == x_src_filename
    assert sheet_ref.src_sheet_name == bk_sheet_name
    assert sheet_ref.src_ii_bk_type is None
    assert sheet_ref.idea_type_exists is None
    assert sheet_ref.src_idea_type is None
    assert sheet_ref.dst_brick_type is None
    assert sheet_ref.dst_sheet_name is None
    assert set(sheet_ref.__dict__.keys()) == {
        "src_filename",
        "src_sheet_name",
        "idea_type_exists",
        "src_ii_bk_type",
        "src_idea_type",
        "dst_brick_type",
        "dst_sheet_name",
    }


def test_set_src_ii_bk_type_SetsAttr_Scenario0_ReturnsIiMatchWhenIiExists():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", src_sheet_name="abc_ii12300_test")
    assert not sheet_ref.src_ii_bk_type
    # WHEN
    sheet_ref.set_src_ii_bk_type()
    # THEN
    assert sheet_ref.src_ii_bk_type == "ii12300"


def test_set_src_ii_bk_type_SetsAttr_Scenario1_ReturnsBkMatchWhenBkExistsAndNoIiExists():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", src_sheet_name="abc_bk45600_test")
    assert not sheet_ref.src_ii_bk_type
    # WHEN
    sheet_ref.set_src_ii_bk_type()
    # THEN
    assert sheet_ref.src_ii_bk_type == "bk45600"


def test_set_src_ii_bk_type_SetsAttr_Scenario2_ReturnsIiMatchWhenBothIiAndBkExist():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", src_sheet_name="bk456_middle_ii12300_end")
    assert not sheet_ref.src_ii_bk_type
    # WHEN
    sheet_ref.set_src_ii_bk_type()
    # THEN
    assert sheet_ref.src_ii_bk_type == "ii12300"


def test_set_src_ii_bk_type_SetsAttr_Scenario3_ReturnsNoneWhenNoMatchExists():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", src_sheet_name="abc_xyz_test")
    assert not sheet_ref.src_ii_bk_type
    # WHEN
    sheet_ref.set_src_ii_bk_type()
    # THEN
    assert sheet_ref.src_ii_bk_type is None


def test_set_idea_type_exists_SetsAttr_Scenario0_SetTrueWhenExists():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "ii00122_end")
    sheet_ref.set_src_ii_bk_type()
    idea_config = get_idea_config_dict()
    assert sheet_ref.idea_type_exists is None
    assert not sheet_ref.src_idea_type
    # WHEN
    sheet_ref.set_idea_type_exists(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is True
    assert sheet_ref.src_idea_type == "ii00122"


def test_set_idea_type_exists_SetsAttr_Scenario2_SetFalse():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "ii123_end")
    sheet_ref.set_src_ii_bk_type()
    idea_config = get_idea_config_dict()
    assert sheet_ref.idea_type_exists is None
    assert not sheet_ref.src_idea_type
    # WHEN
    sheet_ref.set_idea_type_exists(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is False
    assert not sheet_ref.src_idea_type


def test_set_idea_type_exists_SetsAttr_Scenario3_bk00120SetTrue():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "bk00123ii_x")
    sheet_ref.set_src_ii_bk_type()
    idea_config = get_idea_config_dict()
    assert sheet_ref.idea_type_exists is None
    assert not sheet_ref.src_idea_type
    # WHEN
    sheet_ref.set_idea_type_exists(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is True
    assert sheet_ref.src_idea_type == "ii00123"


def test_set_idea_type_exists_SetsAttr_Scenario0_SetFalse():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "ii123_end")
    sheet_ref.set_src_ii_bk_type()
    assert not sheet_ref.src_ii_bk_type
    idea_config = get_idea_config_dict()
    assert not sheet_ref.idea_type_exists
    assert not sheet_ref.src_idea_type
    # WHEN
    sheet_ref.set_idea_type_exists(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is False
    assert not sheet_ref.src_idea_type


def test_set_idea_type_exists_SetsAttr_Scenario1_SetFalse():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "bk00123ii_x")
    sheet_ref.set_src_ii_bk_type()
    idea_config = get_idea_config_dict()
    assert sheet_ref.idea_type_exists is None
    assert not sheet_ref.src_idea_type
    # WHEN
    sheet_ref.set_idea_type_exists(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is True
    assert sheet_ref.src_idea_type == "ii00123"


def test_set_dst_attrs_SetsAttr_Scenario0_bk00120SetTrue():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "bk00123ii_x")
    idea_config = get_idea_config_dict()
    assert sheet_ref.idea_type_exists is None
    assert not sheet_ref.src_idea_type
    assert not sheet_ref.dst_brick_type
    # WHEN
    sheet_ref.set_dst_attrs(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is True
    assert sheet_ref.src_idea_type == "ii00123"
    assert sheet_ref.dst_brick_type == "bk00123"


def test_set_dst_attrs_SetsAttr_Scenario0_Set_False():
    # ESTABLISH
    sheet_ref = SheetRef("test.xlsx", "bk001ii123_end23ii_x")
    idea_config = get_idea_config_dict()
    assert not sheet_ref.idea_type_exists
    assert not sheet_ref.src_idea_type
    assert not sheet_ref.dst_brick_type
    # WHEN
    sheet_ref.set_dst_attrs(idea_config)
    # THEN
    assert sheet_ref.idea_type_exists is False
    assert not sheet_ref.src_idea_type
    assert not sheet_ref.dst_brick_type


def create_df(**kwargs):
    return pandas_DataFrame({**kwargs})


def test_validate_idea_columns_ReturnsDf_Scenario01_ReturnsDataframeWhenAllColumnsPresent():
    # ESTABLISH
    df = create_df(moment_rope=[";mmt01;"], person_name=["Alice"])
    config = {"src_columns": ["moment_rope", "person_name"]}
    # WHEN
    result = validate_idea_columns(df, config, strict=True)
    # THEN
    assert result is not None


def test_validate_idea_columns_ReturnsNone_Scenario02_ReturnsNoneWhenColumnsMissingAndNotStrict():
    # ESTABLISH
    df = create_df(person_name=["Alice"])
    config = {"src_columns": ["moment_rope", "person_name"]}
    # WHEN
    result = validate_idea_columns(df, config, strict=False)
    # THEN
    assert result is None


def test_validate_idea_columns_RaisesValueError_Scenario03_RaisesWhenColumnsMissingAndStrict():
    # ESTABLISH
    df = create_df(person_name=["Alice"])
    config = {"src_columns": ["moment_rope", "person_name"]}
    # WHEN / THEN
    with pytest_raises(ValueError, match="moment_rope"):
        validate_idea_columns(df, config, strict=True)


def test_validate_idea_columns_ReturnsDf_Scenario04_ReturnsDataframeWhenNoSrcColumnsInConfig():
    # ESTABLISH
    df = create_df(person_name=["Alice"])
    config = {}
    # WHEN
    result = validate_idea_columns(df, config, strict=True)
    # THEN
    assert result is not None


def test_ideas_sheets_to_brick_sheets_Scenario0_TwoTuples(tmp_path: Path):
    """Returns one (filename, sheet_name) tuple per brick_type sheet copied."""
    # ESTABLISH
    empty_b_src_dir = tmp_path / "bricks"
    empty_b_src_dir.mkdir()

    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00119_Sales"
    ws1.append(["product", "units", "revenue"])
    ws1.append(["widget", 10, 500])
    ws1.append(["gadget", 5, 250])

    ws2 = wb.create_sheet("ii00112_Costs")
    ws2.append(["category", "amount"])
    ws2.append(["rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    result = ideas_sheets_to_brick_sheets(populated_idea_dir, empty_b_src_dir)
    # THEN
    x_filename = "AllSales.xlsx"
    print(f"{result=}")
    expected_bk00104_sheet_ref = SheetRef(x_filename, "ii00112_Costs")
    expected_bk00120_sheet_ref = SheetRef(x_filename, "ii00119_Sales")
    idea_config = get_idea_config_dict()
    expected_bk00104_sheet_ref.set_dst_attrs(idea_config)
    expected_bk00120_sheet_ref.set_dst_attrs(idea_config)
    assert expected_bk00104_sheet_ref == result[0]
    assert expected_bk00120_sheet_ref == result[1]
    assert len(result) == 2


def test_ideas_sheets_to_brick_sheets_Scenario1_CreatesDestinationFile(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    empty_b_src_dir = tmp_path / "bricks"
    empty_b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, empty_b_src_dir)
    # THEN
    allsales_path = os_path_join(str(empty_b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 1
    assert df["revenue"].sum() == 750


def test_ideas_sheets_to_brick_sheets_Scenario3_DestinationFileHas_spark_num_SetBy_b_src_dir(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    brick_wb = openpyxl_Workbook()
    brick_ws1 = brick_wb.active
    brick_ws1.title = "bk00120_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    brick_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    brick_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    brick_wb.save(b_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, b_src_dir)
    # THEN
    allsales_path = os_path_join(str(b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    assert list(df.columns) == expected_dst_columns
    assert len(df) == 2
    assert df[kw.spark_num].min() == 11
    assert df[kw.spark_num].min() == curr_spark_num + 1
    assert df["revenue"].sum() == 750


def test_ideas_sheets_to_brick_sheets_Scenario4_ParameterSparkNumAccepted(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()
    populated_idea_dir = tmp_path / kw.idea
    populated_idea_dir.mkdir()
    brick_wb = openpyxl_Workbook()
    brick_ws1 = brick_wb.active
    brick_ws1.title = "bk00120_Sales"
    expected_dst_columns = [kw.spark_num, kw.spark_face, "product", "units", "revenue"]
    brick_ws1.append(expected_dst_columns)
    curr_spark_num = 10
    brick_ws1.append([curr_spark_num, exx.sue, "widget", 10, 500])
    brick_wb.save(b_src_dir / "OtherFile.xlsx")

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    ws1.append([exx.sue, "gadget", 5, 250])

    ws2 = wb.create_sheet("ii00104_Costs")
    ws2.append([kw.spark_face, "category", "amount"])
    ws2.append([exx.sue, "rent", 1000])
    wb.create_sheet("Summary")  # non-BR, should be ignored
    wb.save(populated_idea_dir / "AllSales.xlsx")
    db_max_spark_num = 22

    # WHEN
    ideas_sheets_to_brick_sheets(populated_idea_dir, b_src_dir, db_max_spark_num)
    # THEN
    allsales_path = os_path_join(str(b_src_dir), "AllSales.xlsx")
    df = pandas_read_excel(allsales_path, sheet_name="bk00120_Sales")
    assert df[kw.spark_num].min() != 11
    assert df[kw.spark_num].min() != curr_spark_num + 1
    assert df[kw.spark_num].min() == db_max_spark_num + 1


def test_ideas_sheets_to_brick_sheets_Scenario5_src_dir_IsEmptied(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    idea_dir = tmp_path / kw.idea
    idea_dir.mkdir()
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "bk00120_Sales"
    ws1.append([kw.spark_face, "product", "units", "revenue"])
    ws1.append([exx.sue, "widget", 10, 500])
    wb.save(idea_dir / "AllSales.xlsx")
    assert count_dirs_files(idea_dir) == 1

    # WHEN
    ideas_sheets_to_brick_sheets(idea_dir, b_src_dir)
    # THEN
    assert count_dirs_files(idea_dir) == 0


def test_ideas_sheets_to_brick_sheets_Scenario6_src_num_Exists(
    tmp_path: Path,
):
    """Each copied sheet can be read by pandas and contains the original data."""
    # ESTABLISH
    idea_dir = tmp_path / kw.idea
    idea_dir.mkdir()
    b_src_dir = tmp_path / "bricks"
    b_src_dir.mkdir()

    wb = openpyxl_Workbook()
    ws1 = wb.active
    ws1.title = "ii00120_Sales"
    ws1.append([kw.spark_num, kw.spark_face, "product", "units", "revenue"])
    ws1.append(["", exx.sue, "widget", 10, 500])
    idea_allsales_path = idea_dir / "AllSales.xlsx"
    wb.save(idea_allsales_path)
    assert count_dirs_files(idea_dir) == 1
    assert count_dirs_files(b_src_dir) == 0

    # WHEN
    ideas_sheets_to_brick_sheets(idea_dir, b_src_dir)
    # THEN
    assert count_dirs_files(idea_dir) == 0
    assert count_dirs_files(b_src_dir) == 1
    brick_allsales_path = b_src_dir / "AllSales.xlsx"
    df = pandas_read_excel(brick_allsales_path, sheet_name="bk00120_Sales")
    assert df[kw.spark_num].min() == 1
