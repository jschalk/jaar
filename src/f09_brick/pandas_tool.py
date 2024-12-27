from src.f00_instrument.file import (
    set_dir,
    save_file,
    create_path,
    get_dir_filenames,
    get_dir_file_strs,
    open_file,
)
from src.f00_instrument.db_toolbox import get_grouping_with_all_values_equal_sql_query
from src.f08_pidgin.map import MapCore
from src.f08_pidgin.pidgin import (
    PidginUnit,
    pidginable_atom_args,
    get_pidginunit_from_json,
)
from src.f08_pidgin.pidgin_config import get_pidgin_args_jaar_types
from src.f09_brick.brick_config import (
    get_brick_elements_sort_order,
    get_brick_category_ref,
    get_brick_format_filename,
)
from numpy import float64, nan as numpy_nan
from pandas import (
    DataFrame,
    read_csv as pandas_read_csv,
    read_sql_query as pandas_read_sql_query,
    ExcelWriter,
    read_excel as pandas_read_excel,
)
from openpyxl import load_workbook as openpyxl_load_workbook
from sqlite3 import connect as sqlite3_connect
from os.path import exists as os_path_exists, dirname as os_path_dirname


def save_dataframe_to_csv(x_df: DataFrame, x_dir: str, x_filename: str):
    save_file(x_dir, x_filename, get_ordered_csv(x_df))


def get_sorting_columns(
    existing_columns: set[str], sorting_columns: list[str] = None
) -> list[str]:
    if sorting_columns is None:
        sorting_columns = get_brick_elements_sort_order()
    sort_columns_in_existing = set(sorting_columns).intersection(existing_columns)
    return [
        sort_col for sort_col in sorting_columns if sort_col in sort_columns_in_existing
    ]


def get_ordered_csv(x_df: DataFrame, sorting_columns: list[str] = None) -> str:
    new_sorting_columns = get_sorting_columns(set(x_df.columns), sorting_columns)
    x_df.sort_values(new_sorting_columns, inplace=True)
    x_df.reset_index(inplace=True)
    x_df.drop(columns=["index"], inplace=True)
    return x_df.to_csv(index=False).replace("\r", "")


def open_csv(x_file_dir: str, x_filename: str) -> DataFrame:
    return pandas_read_csv(create_path(x_file_dir, x_filename))


def get_sheet_names(x_path: str) -> list[str]:
    return openpyxl_load_workbook(x_path).sheetnames


def get_all_excel_sheet_names(
    dir: str, sub_strs: set[str] = None
) -> set[(str, str, str)]:
    if sub_strs is None:
        sub_strs = set()
    excel_files = get_dir_filenames(dir, {"xlsx"})
    sheet_names = set()
    for relative_dir, filename in excel_files:
        absolute_dir = create_path(dir, relative_dir)
        absolute_path = create_path(absolute_dir, filename)
        file_sheet_names = get_sheet_names(absolute_path)
        for sheet_name in file_sheet_names:
            if not sub_strs:
                sheet_names.add((absolute_dir, filename, sheet_name))
            else:
                for sub_str in sub_strs:
                    if sheet_name.find(sub_str) >= 0:
                        sheet_names.add((absolute_dir, filename, sheet_name))
    return sheet_names


def get_relevant_columns_dataframe(
    src_df: DataFrame,
    relevant_columns: list[str] = None,
    relevant_columns_necessary: bool = True,
) -> DataFrame:
    if relevant_columns is None:
        relevant_columns = get_brick_elements_sort_order()
    current_columns = set(src_df.columns.to_list())
    relevant_columns_set = set(relevant_columns)
    current_relevant_columns = current_columns.intersection(relevant_columns_set)
    relevant_cols_in_order = [
        r_col for r_col in relevant_columns if r_col in current_relevant_columns
    ]
    return src_df[relevant_cols_in_order]


def boat_staging_str():
    return "boat_staging"


def boat_agg_str():
    return "boat_agg"


def boat_valid_str():
    return "boat_valid"


def get_boat_staging_grouping_with_all_values_equal_df(
    x_df: DataFrame, group_by_list: list
) -> DataFrame:
    df_columns = set(x_df.columns)
    grouping_columns = get_sorting_columns(df_columns, group_by_list)
    value_columns = df_columns.difference(grouping_columns)

    if grouping_columns == []:
        return x_df
    with sqlite3_connect(":memory:") as conn:
        x_df.to_sql(boat_staging_str(), conn, index=False)
        query_str = get_grouping_with_all_values_equal_sql_query(
            x_table=boat_staging_str(),
            group_by_columns=grouping_columns,
            value_columns=value_columns,
        )
        return pandas_read_sql_query(query_str, conn)


def get_dataframe_pidginable_columns(x_df: DataFrame) -> set[str]:
    return {x_column for x_column in x_df.columns if x_column in pidginable_atom_args()}


def translate_single_column_dataframe(
    x_df: DataFrame, x_mapunit: MapCore, column_name: str
) -> DataFrame:
    if column_name in x_df:
        row_count = len(x_df)
        for cur_row in range(row_count):
            otx_value = x_df.iloc[cur_row][column_name]
            inx_value = x_mapunit.reveal_inx(otx_value)
            x_df.at[cur_row, column_name] = inx_value
    return x_df


def translate_all_columns_dataframe(x_df: DataFrame, x_pidginunit: PidginUnit):
    if x_pidginunit is None:
        return None

    column_names = set(x_df.columns)
    pidginable_columns = column_names.intersection(pidginable_atom_args())
    for pidginable_column in pidginable_columns:
        jaar_type = get_pidgin_args_jaar_types().get(pidginable_column)
        x_mapunit = x_pidginunit.get_mapunit(jaar_type)
        translate_single_column_dataframe(x_df, x_mapunit, pidginable_column)


def move_otx_csvs_to_pidgin_inx(face_dir: str):
    bow_dir = create_path(face_dir, "bow")
    aft_dir = create_path(face_dir, "aft")
    pidgin_filename = "pidgin.json"
    pidginunit_json = open_file(face_dir, pidgin_filename)
    face_pidginunit = get_pidginunit_from_json(pidginunit_json)
    bow_dir_files = get_dir_file_strs(bow_dir, delete_extensions=False)
    for x_file_name in bow_dir_files.keys():
        x_df = open_csv(bow_dir, x_file_name)
        translate_all_columns_dataframe(x_df, face_pidginunit)
        save_dataframe_to_csv(x_df, aft_dir, x_file_name)


def _get_pidgen_brick_format_filenames() -> set[str]:
    brick_numbers = set(get_brick_category_ref().get("map_acct_id"))
    brick_numbers.update(set(get_brick_category_ref().get("map_group_id")))
    brick_numbers.update(set(get_brick_category_ref().get("map_idea")))
    brick_numbers.update(set(get_brick_category_ref().get("map_road")))
    return {f"{brick_number}.xlsx" for brick_number in brick_numbers}


def _get_deal_brick_format_filenames() -> set[str]:
    brick_numbers = set(get_brick_category_ref().get("dealunit"))
    brick_numbers.update(set(get_brick_category_ref().get("deal_cashbook")))
    brick_numbers.update(set(get_brick_category_ref().get("deal_purview_episode")))
    brick_numbers.update(set(get_brick_category_ref().get("deal_timeline_hour")))
    brick_numbers.update(set(get_brick_category_ref().get("deal_timeline_month")))
    brick_numbers.update(set(get_brick_category_ref().get("deal_timeline_weekday")))
    return {f"{brick_number}.xlsx" for brick_number in brick_numbers}


def append_df_to_excel(file_path: str, sheet_name: str, dataframe: DataFrame):
    try:
        # Load the existing workbook
        workbook = openpyxl_load_workbook(file_path)

        # Check if the sheet exists, if not create it
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
            sheet = workbook[sheet_name]
            # Add column names to the new sheet
            for col_num, column_title in enumerate(dataframe.columns, 1):
                sheet.cell(row=1, column=col_num, value=column_title)
            start_row = 2  # Start appending data from the second row
        else:
            sheet = workbook[sheet_name]
            start_row = sheet.max_row + 1

        # Convert the DataFrame to a list of rows
        rows = dataframe.to_dict(orient="split")["data"]

        # Append the rows to the sheet
        for i, row in enumerate(rows, start_row):
            for j, value in enumerate(row, 1):  # 1-based index for Excel
                sheet.cell(row=i, column=j, value=value)

        # Save changes to the workbook
        workbook.save(file_path)
        # prt("Data appended successfully!")

    except FileNotFoundError:
        # If the file doesn't exist, create a new one
        # prt(f"{file_path} not found. Creating a new Excel file.")
        dataframe.to_excel(file_path, index=False, sheet_name=sheet_name)


class pandas_tools_ExcelWriterException(Exception):
    pass


def upsert_sheet(
    file_path: str, sheet_name: str, dataframe: DataFrame, replace: bool = False
):
    # sourcery skip: remove-redundant-exception, simplify-single-exception-tuple
    set_dir(os_path_dirname(file_path))
    """
    Updates or creates an Excel sheet with a specified DataFrame.

    Args:
    - file_path (str): The path to the Excel file.
    - sheet_name (str): The name of the sheet to update or create.
    - dataframe (pd.DataFrame): The DataFrame to write to the sheet.
    """
    # Check if the file exists
    if not os_path_exists(file_path):
        # If file does not exist, create it with the specified sheet
        with ExcelWriter(file_path, engine="xlsxwriter") as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        return

    # If the file exists, check for the sheet
    try:
        if replace:
            with ExcelWriter(
                file_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
            ) as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            append_df_to_excel(file_path, sheet_name, dataframe)

    except (PermissionError, FileNotFoundError, OSError) as e:
        raise pandas_tools_ExcelWriterException(f"An error occurred: {e}") from e


def sheet_exists(file_path: str, sheet_name: str):
    """
    Checks if a specific sheet exists in an Excel workbook.

    Args:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to check.

    Returns:
        bool: True if the sheet exists, False otherwise.
    """
    if not os_path_exists(file_path):
        return False

    try:
        return sheet_name in set(get_sheet_names(file_path))
    except Exception as e:
        return False


def split_excel_into_dirs(
    input_file: str, output_dir: str, column_name: str, file_name: str, sheet_name: str
):
    """
    Splits an Excel file into multiple Excel files, each containing rows
    corresponding to a unique value in the specified column.

    Args:
        input_file (str): Path to the input Excel file.
        output_dir (str): Directory where the output files will be saved.
        column_name (str): Column to split by unique values.
    """
    # Create the output directory if it doesn't exist
    set_dir(output_dir)
    df = pandas_read_excel(input_file, sheet_name=sheet_name)

    # Check if the column exists
    if column_name not in df.columns:
        raise ValueError(f"Column '{column_name}' does not exist in the input file.")

    # Group by unique values in the column
    unique_values = df[column_name].unique()

    for value in unique_values:
        if float64 != type(value):
            # Filter rows for the current unique value
            filtered_df = df[df[column_name] == value]

            # Create a safe subdirectory name for the unique value
            safe_value = str(value).replace("/", "_").replace("\\", "_")
            subdirectory = create_path(output_dir, safe_value)
            # Create the subdirectory if it doesn't exist
            set_dir(subdirectory)
            # Define the output file path
            output_file = create_path(subdirectory, f"{file_name}.xlsx")
            upsert_sheet(output_file, sheet_name, filtered_df)


def if_nan_return_None(x_obj: any) -> any:
    # sourcery skip: equality-identity, remove-redundant-if
    # If the value is NaN, the comparison value != value
    return None if x_obj != x_obj else x_obj
