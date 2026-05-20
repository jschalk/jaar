from ch00_py.file_toolbox import create_path, open_json
from ch17_brick.brick_db_tool import (
    EXCEL_READER_CONFIG,
    _normalize_excel_value,
    create_brick_df_from_file,
    get_excel_reader_config_path,
    get_excel_reader_src_config,
)
from ch99_glossary.ch_keyword import Ch17Keywords as kw
from openpyxl import Workbook


def test_get_excel_reader_config_path_ReturnsObj() -> str:
    # ESTABLISH
    chapter_dir = create_path("src", "ch17_brick")
    # WHEN / THEN
    assert get_excel_reader_config_path() == create_path(
        chapter_dir, "excel_reader.json"
    )


def test_get_excel_reader_src_config_ReturnsObj_Brick() -> str:
    # ESTABLISH / WHEN
    excel_reader_json_config = get_excel_reader_src_config()

    # THEN
    assert excel_reader_json_config
    expected_dict = open_json(get_excel_reader_config_path())
    assert excel_reader_json_config == expected_dict
    assert set(excel_reader_json_config.keys()) == {
        "normalization_rules",
        "canonical_cell_types",
        "error_handling",
    }


def test_EXCEL_READER_CONFIG_Exists():
    # ESTABLISH
    expected_config = {
        "canonical_cell_types": {
            "TEXT": "str",
            "INT": "int",
            "REAL": "float",
            "EMPTY": None,
        },
        "normalization_rules": [
            {
                "rule_name": "nan_to_none",
                "match_type": "is_nan",
                "replacement_value": None,
            },
            {
                "rule_name": "empty_string_to_none",
                "match_type": "exact_string",
                "match_value": "",
                "replacement_value": None,
            },
            {
                "rule_name": "whitespace_string_to_none",
                "match_type": "whitespace_string",
                "replacement_value": None,
            },
            {
                "rule_name": "excel_true_formula_to_text",
                "match_type": "exact_string",
                "match_value": "=TRUE()",
                "replacement_value": 1,
            },
            {
                "rule_name": "excel_false_formula_to_text",
                "match_type": "exact_string",
                "match_value": "=FALSE()",
                "replacement_value": 0,
            },
            # {
            #     "rule_name": "python_true_to_text",
            #     "match_type": "python_bool",
            #     "match_value": True,
            #     "replacement_value": "TRUE",
            # },
            # {
            #     "rule_name": "python_false_to_text",
            #     "match_type": "python_bool",
            #     "match_value": False,
            #     "replacement_value": "FALSE",
            # },
        ],
        "error_handling": {
            "on_conversion_error": "collect",
            "error_column_name": "error_message",
        },
    }
    # WHEN / THEN
    assert EXCEL_READER_CONFIG == expected_config


def test_normalize_excel_value_ReturnsObj_Scenario0_Converts_TRUE_Formula():
    # ESTABLISH / WHEN
    result = _normalize_excel_value("=TRUE()")
    # THEN
    assert result == 1


def test_normalize_excel_value_ReturnsObj_Scenario1_Converts_FALSE_Formula():
    # ESTABLISH / WHEN
    result = _normalize_excel_value("=FALSE()")
    # THEN
    assert result == 0


def test_normalize_excel_value_ReturnsObj_Scenario2_Converts_True_Boolean():
    # ESTABLISH / WHEN
    result = _normalize_excel_value(True)
    # THEN
    assert result == True


def test_normalize_excel_value_ReturnsObj_Scenario3_Converts_False_Boolean():
    # ESTABLISH / WHEN
    result = _normalize_excel_value(False)
    # THEN
    assert result == False


def test_normalize_excel_value_ReturnsObj_Scenario4_Converts_EmptyString_ToNone():
    # ESTABLISH / WHEN
    result = _normalize_excel_value("")
    # THEN
    assert result is None


def test_normalize_excel_value_ReturnsObj_Scenario5_Converts_WhitespaceString_ToNone():
    # ESTABLISH / WHEN
    result = _normalize_excel_value("   ")
    # THEN
    assert result is None


def test_normalize_excel_value_ReturnsObj_Scenario6_Preserves_NormalString():
    # ESTABLISH / WHEN
    result = _normalize_excel_value("Sue")
    # THEN
    assert result == "Sue"


def test_normalize_excel_value_ReturnsObj_Scenario7_Preserves_Integer():
    # ESTABLISH / WHEN
    result = _normalize_excel_value(55)
    # THEN
    assert result == 55


def test_normalize_excel_value_ReturnsObj_Scenario8_Preserves_Float():
    # ESTABLISH / WHEN
    result = _normalize_excel_value(55.5)
    # THEN
    assert result == 55.5


def test_normalize_excel_value_ReturnsObj_Scenario9_Converts_None_ToNone():
    # ESTABLISH / WHEN
    result = _normalize_excel_value(None)
    # THEN
    assert result is None


def test_create_brick_df_from_file_ReturnsObj_Scenario0_Converts_TRUE_FALSE_Formulas(
    temp3_fs,
):
    # ESTABLISH
    excel_path = create_path(str(temp3_fs), "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["flag"])
    ws.append(["=TRUE()"])
    ws.append(["=FALSE()"])
    wb.save(excel_path)

    # WHEN
    df = create_brick_df_from_file(excel_path, "sheet1")

    # THEN
    assert df["flag"].iloc[0] == 1
    assert df["flag"].iloc[1] == 0


def test_create_brick_df_from_file_ReturnsObj_Scenario1_Converts_Python_Booleans(
    temp3_fs,
):
    # ESTABLISH
    excel_path = create_path(str(temp3_fs), "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["flag"])
    ws.append([True])
    ws.append([False])
    wb.save(excel_path)

    # WHEN
    df = create_brick_df_from_file(excel_path, "sheet1")

    # THEN
    assert df["flag"].iloc[0] == True
    assert df["flag"].iloc[1] == False


def test_create_brick_df_from_file_ReturnsObj_Scenario2_Converts_Empty_And_Whitespace_ToNone(
    temp3_fs,
):
    # ESTABLISH
    excel_path = create_path(str(temp3_fs), "test.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"
    ws.append(["value"])
    ws.append([""])
    ws.append(["   "])
    wb.save(excel_path)

    # WHEN
    df = create_brick_df_from_file(excel_path, "sheet1")

    # THEN
    assert df["value"].iloc[0] is None
    assert df["value"].iloc[1] is None


def test_create_brick_df_from_file_ReturnsObj_Scenario3_Preserves_Normal_Data(
    temp3_fs,
):
    # ESTABLISH
    excel_path = create_path(str(temp3_fs), "test.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "sheet1"

    ws.append(["id", "name", "value"])
    ws.append([1, "Sue", 10])
    ws.append([2, "Bob", 20.5])

    wb.save(excel_path)

    # WHEN
    df = create_brick_df_from_file(excel_path, "sheet1")

    # THEN
    assert df["id"].iloc[0] == 1
    assert df["name"].iloc[0] == "Sue"
    assert df["value"].iloc[1] == 20.5
