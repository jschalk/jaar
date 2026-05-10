from ch00_py.file_toolbox import create_path
from ch17_brick.brick_db_tool import normalize_excel_file_for_loading
from openpyxl import Workbook, load_workbook


def test_normalize_excel_file_for_loading_ReturnsObj_Scenario0_Converts_TRUE_FALSE(
    temp3_fs,
):
    # ESTABLISH an Excel file written into temp filesystem
    ex_filename = "test_true_false.xlsx"
    file_path = create_path(str(temp3_fs), ex_filename)

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "=TRUE()"
    ws["A2"] = "=FALSE()"
    ws["A3"] = "normal_value"

    wb.save(file_path)

    # WHEN normalization is applied
    normalize_excel_file_for_loading(file_path)

    # THEN values are converted in-place
    wb2 = load_workbook(file_path)
    ws2 = wb2.active

    assert ws2["A1"].value == "TRUE"
    assert ws2["A2"].value == "FALSE"
    assert ws2["A3"].value == "normal_value"


def test_normalize_excel_file_for_loading_ReturnsObj_Scenario1_Converts_Python_Booleans(
    temp3_fs,
):
    # ESTABLISH an Excel file with actual Python boolean values
    ex_filename = "test_bool_values.xlsx"
    file_path = create_path(str(temp3_fs), ex_filename)

    wb = Workbook()
    ws = wb.active

    ws["A1"] = True
    ws["A2"] = False
    ws["A3"] = "TRUE"

    wb.save(file_path)

    # WHEN normalization is applied
    normalize_excel_file_for_loading(file_path)

    # THEN booleans are converted to text
    wb2 = load_workbook(file_path)
    ws2 = wb2.active

    assert ws2["A1"].value == "TRUE"
    assert ws2["A2"].value == "FALSE"
    assert ws2["A3"].value == "TRUE"


def test_normalize_excel_file_for_loading_ReturnsObj_Scenario2_DoesNotModify_UnrelatedValues(
    temp3_fs,
):
    # ESTABLISH an Excel file with mixed unrelated values
    ex_filename = "test_no_side_effects.xlsx"
    file_path = create_path(str(temp3_fs), ex_filename)

    wb = Workbook()
    ws = wb.active

    ws["A1"] = "hello"
    ws["A2"] = 123
    ws["A3"] = None
    ws["A4"] = "=TRUE()"

    wb.save(file_path)

    # WHEN normalization is applied
    normalize_excel_file_for_loading(file_path)

    # THEN only target patterns are changed
    wb2 = load_workbook(file_path)
    ws2 = wb2.active

    assert ws2["A1"].value == "hello"
    assert ws2["A2"].value == 123
    assert ws2["A3"].value is None
    assert ws2["A4"].value == "TRUE"
