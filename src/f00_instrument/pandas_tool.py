from src.f00_instrument.file import save_file, create_file_path, get_all_filenames
from pandas import DataFrame, read_csv as pandas_read_csv
from openpyxl import load_workbook as openpyxl_load_workbook


def get_sorting_priority_column_headers() -> list[str]:
    return [
        "face_id",
        "python_type",
        "owner_id",
        "acct_id",
        "group_id",
        "parent_road",
        "label",
        "road",
        "base",
        "need",
        "pick",
        "team_id",
        "awardee_id",
        "healer_id",
        "numor",
        "denom",
        "addin",
        "base_item_active_requisite",
        "begin",
        "close",
        "credit_belief",
        "debtit_belief",
        "credit_vote",
        "debtit_vote",
        "credor_respect",
        "debtor_respect",
        "fopen",
        "fnigh",
        "fund_pool",
        "give_force",
        "gogo_want",
        "mass",
        "max_tree_traverse",
        "morph",
        "nigh",
        "open",
        "divisor",
        "pledge",
        "problem_bool",
        "purview_timestamp",
        "stop_want",
        "take_force",
        "tally",
        "fund_coin",
        "penny",
        "respect_bit",
        "otx_road_delimiter",
        "inx_road_delimiter",
        "unknown_word",
        "otx_word",
        "inx_word",
        "otx_label",
        "inx_label",
    ]


def save_dataframe_to_csv(x_dt: DataFrame, x_dir: str, x_filename: str):
    save_file(x_dir, x_filename, get_ordered_csv(x_dt))


def get_ordered_csv(x_dt: DataFrame, sorting_columns: list[str] = None) -> str:
    if sorting_columns is None:
        sorting_columns = get_sorting_priority_column_headers()
    sort_columns_in_dt = set(sorting_columns).intersection(set(x_dt.columns))
    new_sorting_columns = [
        sort_col for sort_col in sorting_columns if sort_col in sort_columns_in_dt
    ]
    x_dt.sort_values(new_sorting_columns, inplace=True)
    x_dt.reset_index(inplace=True)
    x_dt.drop(columns=["index"], inplace=True)
    return x_dt.to_csv(index=False).replace("\r", "")


def open_csv(x_file_dir: str, x_filename: str) -> DataFrame:
    return pandas_read_csv(create_file_path(x_file_dir, x_filename))


def get_all_excel_sheetnames(dir: str) -> set[(str, str, str)]:
    excel_files = get_all_filenames(dir, {"xlsx"})
    print(f"{excel_files=}")
    sheet_names = set()
    for relative_dir, filename in excel_files:
        absolute_dir = create_file_path(dir, relative_dir)
        absolute_path = create_file_path(absolute_dir, filename)
        print(f"{relative_dir=} {filename=}")
        print(f"{absolute_path=} ")
        file_sheet_names = openpyxl_load_workbook(absolute_path).sheetnames
        for sheet_name in file_sheet_names:
            sheet_names.add((absolute_dir, filename, sheet_name))
    print(f"{sheet_names=}")
    return sheet_names
