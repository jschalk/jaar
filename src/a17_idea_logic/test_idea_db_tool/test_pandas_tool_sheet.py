from src.a00_data_toolbox.file_toolbox import create_path, set_dir
from src.a17_idea_logic._utils.env_a17 import (
    env_dir_setup_cleanup,
    idea_fisc_mstr_dir,
)
from src.a17_idea_logic.idea_db_tool import (
    sheet_exists,
    upsert_sheet,
    get_all_excel_sheet_names,
    split_excel_into_dirs,
    if_nan_return_None,
    append_df_to_excel,
    set_dataframe_first_two_columns,
    check_dataframe_column_names,
    update_all_face_name_event_int_columns,
)
from pytest import fixture as pytest_fixture, raises as pytest_raises
from pandas import DataFrame, read_excel as pandas_read_excel
from pandas.testing import assert_frame_equal as pandas_testing_assert_frame_equal
from openpyxl import (
    load_workbook as openpyxl_load_workbook,
    Workbook as openpyxl_Workbook,
)
from os.path import exists as os_path_exists
from numpy import nan as numpy_nan, float64


def test_append_df_to_excel_CreatesSheet(env_dir_setup_cleanup):
    # ESTABLISH
    test_file = create_path(idea_fisc_mstr_dir(), "test.xlsx")
    append_data = {
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "City": ["New York", "Los Angeles"],
    }
    append_df = DataFrame(append_data)
    set_dir(idea_fisc_mstr_dir())
    assert os_path_exists(test_file) is False

    # WHEN
    append_df_to_excel(file_path=test_file, sheet_name="Sheet1", dataframe=append_df)

    # THEN
    assert os_path_exists(test_file)
    workbook = openpyxl_load_workbook(test_file)
    sheet = workbook["Sheet1"]
    rows = list(sheet.iter_rows(values_only=True))
    expected_rows = [
        ("Name", "Age", "City"),
        ("Alice", 25, "New York"),
        ("Bob", 30, "Los Angeles"),
    ]
    assert rows == expected_rows


def test_append_df_to_excel_AppendsToSheet(env_dir_setup_cleanup):
    # ESTABLISH
    set_dir(idea_fisc_mstr_dir())
    test_file = create_path(idea_fisc_mstr_dir(), "test.xlsx")
    initial_data = {
        "Name": ["John", "Doe"],
        "Age": [40, 50],
        "City": ["Boston", "Chicago"],
    }
    append_data = {
        "Name": ["Alice", "Bob"],
        "Age": [25, 30],
        "City": ["New York", "Los Angeles"],
    }
    initial_df = DataFrame(initial_data)
    append_df = DataFrame(append_data)
    initial_df.to_excel(test_file, index=False, sheet_name="Sheet1")

    # WHEN
    append_df_to_excel(file_path=test_file, sheet_name="Sheet1", dataframe=append_df)

    # THEN
    workbook = openpyxl_load_workbook(test_file)
    sheet = workbook["Sheet1"]
    rows = list(sheet.iter_rows(values_only=True))
    expected_rows = [
        ("Name", "Age", "City"),
        ("John", 40, "Boston"),
        ("Doe", 50, "Chicago"),
        ("Alice", 25, "New York"),
        ("Bob", 30, "Los Angeles"),
    ]
    assert rows == expected_rows


@pytest_fixture
def sample_dataframe():
    """Fixture to provide a sample DataFrame."""
    data = {
        "Name": ["Alice", "Bob", "Charlie"],
        "Age": [25, 30, 35],
        "City": ["New York", "Los Angeles", "Chicago"],
    }
    return DataFrame(data)


@pytest_fixture
def temp_excel_file(tmp_path):
    """Fixture to provide a temporary Excel file path."""
    return tmp_path / "test_excel.xlsx"


def test_upsert_sheet_CreatesNewFile(temp_excel_file, sample_dataframe):
    """Test creating a new Excel file with a specified sheet."""
    upsert_sheet(temp_excel_file, "Sheet1", sample_dataframe)
    assert os_path_exists(temp_excel_file)

    # Verify the content of the sheet
    df_read = pandas_read_excel(temp_excel_file, sheet_name="Sheet1")
    assert list(df_read.columns) != []
    pandas_testing_assert_frame_equal(df_read, sample_dataframe)


def test_upsert_sheet_ReplacesExistingSheet(temp_excel_file, sample_dataframe):
    """Test replacing an existing sheet in the Excel file."""
    # ESTABLISH Create the file and write initial data
    initial_data = DataFrame({"A": [1, 2, 3]})
    upsert_sheet(temp_excel_file, "Sheet1", initial_data)

    # WHEN Replace the sheet with new data
    upsert_sheet(temp_excel_file, "Sheet1", sample_dataframe, replace=True)

    # THEN Verify the content of the replaced sheet
    df_read = pandas_read_excel(temp_excel_file, sheet_name="Sheet1")
    pandas_testing_assert_frame_equal(df_read, sample_dataframe)


def test_upsert_sheet_AddNewSheetToExistingFile(temp_excel_file, sample_dataframe):
    """Test adding a new sheet to an existing Excel file."""
    # ESTABLISH Create the file and write initial data to one sheet
    initial_data = DataFrame({"A": [1, 2, 3]})
    upsert_sheet(temp_excel_file, "InitialSheet", initial_data)

    # WHEN Add a new sheet with different data
    upsert_sheet(temp_excel_file, "NewSheet", sample_dataframe, replace=True)

    # THEN Verify both sheets exist and have correct data
    df_initial = pandas_read_excel(temp_excel_file, sheet_name="InitialSheet")
    df_new = pandas_read_excel(temp_excel_file, sheet_name="NewSheet")
    pandas_testing_assert_frame_equal(df_initial, initial_data)
    pandas_testing_assert_frame_equal(df_new, sample_dataframe)


def test_get_all_excel_sheet_names_ReturnsObj_Scenario0_NoPidgin(
    env_dir_setup_cleanup,
):
    # ESTABLISH
    env_dir = idea_fisc_mstr_dir()
    x_dir = create_path(env_dir, "examples_folder")
    ex_filename = "fizzbuzz.xlsx"
    ex_file_path = create_path(x_dir, ex_filename)
    df1 = DataFrame([["AAA", "BBB"]], columns=["spam", "egg"])
    df2 = DataFrame([["ABC", "XYZ"]], columns=["Foo", "Bar"])
    sheet_name1 = "Sheet1x"
    sheet_name2 = "Sheet2x"
    upsert_sheet(ex_file_path, sheet_name1, df1)
    upsert_sheet(ex_file_path, sheet_name2, df2)

    # WHEN
    x_sheet_names = get_all_excel_sheet_names(env_dir)

    # THEN
    assert x_sheet_names
    assert (x_dir, ex_filename, sheet_name1) in x_sheet_names
    assert (x_dir, ex_filename, sheet_name2) in x_sheet_names
    assert len(x_sheet_names) == 2


def test_get_all_excel_sheet_names_ReturnsObj_Scenario1_PidginSheetNames(
    env_dir_setup_cleanup,
):
    # ESTABLISH
    env_dir = idea_fisc_mstr_dir()
    x_dir = create_path(env_dir, "examples_folder")
    ex_filename = "fizzbuzz.xlsx"
    ex_file_path = create_path(x_dir, ex_filename)
    df1 = DataFrame([["AAA", "BBB"]], columns=["spam", "egg"])
    df2 = DataFrame([["ABC", "XYZ"]], columns=["Foo", "Bar"])
    sugar_str = "sugar"
    honey_name1 = "honey1x"
    sugar_name1 = f"{sugar_str}2x"
    sugar_name2 = f"honey_{sugar_str}3x"
    upsert_sheet(ex_file_path, honey_name1, df1)
    upsert_sheet(ex_file_path, sugar_name1, df2)
    upsert_sheet(ex_file_path, sugar_name2, df2)

    # WHEN
    x_sheet_names = get_all_excel_sheet_names(env_dir, sub_strs={sugar_str})

    # THEN
    assert x_sheet_names
    assert (x_dir, ex_filename, honey_name1) not in x_sheet_names
    assert (x_dir, ex_filename, sugar_name1) in x_sheet_names
    assert (x_dir, ex_filename, sugar_name2) in x_sheet_names
    assert len(x_sheet_names) == 2


def test_sheet_exists_ReturnsObj_Scenario1(env_dir_setup_cleanup):
    # ESTABLISH
    env_dir = idea_fisc_mstr_dir()
    x_dir = create_path(env_dir, "examples_folder")
    ex_filename = "fizzbuzz.xlsx"
    ex_file_path = create_path(x_dir, ex_filename)
    df1 = DataFrame([["AAA", "BBB"]], columns=["spam", "egg"])
    sugar_str = "sugar"
    honey_name1 = "honey1x"
    sugar_name1 = f"{sugar_str}2x"
    sugar_name2 = f"honey_{sugar_str}3x"
    assert sheet_exists(ex_file_path, honey_name1) is False
    assert sheet_exists(ex_file_path, sugar_name1) is False
    assert sheet_exists(ex_file_path, sugar_name2) is False

    # WHEN / THEN
    upsert_sheet(ex_file_path, honey_name1, df1)
    assert sheet_exists(ex_file_path, honey_name1)
    assert sheet_exists(ex_file_path, sugar_name1) is False
    assert sheet_exists(ex_file_path, sugar_name2) is False

    # WHEN / THEN
    upsert_sheet(ex_file_path, sugar_name1, df1)
    assert sheet_exists(ex_file_path, honey_name1)
    assert sheet_exists(ex_file_path, sugar_name1)
    assert sheet_exists(ex_file_path, sugar_name2) is False

    # WHEN / THEN
    upsert_sheet(ex_file_path, sugar_name2, df1)
    assert sheet_exists(ex_file_path, honey_name1)
    assert sheet_exists(ex_file_path, sugar_name1)
    assert sheet_exists(ex_file_path, sugar_name2)


@pytest_fixture
def sample_excel_file(tmp_path):
    """Fixture to create a sample Excel file for testing."""
    data = {
        "ID": [1, 2, 3, 4, 5],
        "Dimen": ["A", "B", "A", "C", "B"],
        "Value": [100, 200, 150, 300, 250],
    }
    df = DataFrame(data)
    file_path = tmp_path / "sample.xlsx"
    df.to_excel(file_path, index=False, sheet_name="sheet5")
    return file_path


@pytest_fixture
def output_dir(tmp_path):
    """Fixture to provide a temporary output directory."""
    return tmp_path / "output"


def test_split_excel_into_dirs_RaisesErrorWhenColumnIsInvalid(
    sample_excel_file, output_dir
):
    """Test handling of an invalid column."""
    # ESTABLISH / WHEN / THEN
    with pytest_raises(ValueError, match="Column 'InvalidColumn' does not exist"):
        split_excel_into_dirs(
            sample_excel_file, output_dir, "InvalidColumn", "filename", "sheet5"
        )


def test_split_excel_into_dirs_CreatesFilesWhenColumnIsValid(
    sample_excel_file, output_dir
):
    """Test splitting an Excel file by a valid column."""
    # ESTABLISH
    x_filename = "fizz"
    a_dir = create_path(output_dir, "A")
    b_dir = create_path(output_dir, "B")
    c_dir = create_path(output_dir, "C")
    a_file_path = create_path(a_dir, f"{x_filename}.xlsx")
    b_file_path = create_path(b_dir, f"{x_filename}.xlsx")
    c_file_path = create_path(c_dir, f"{x_filename}.xlsx")
    assert os_path_exists(a_file_path) is False
    assert os_path_exists(b_file_path) is False
    assert os_path_exists(c_file_path) is False

    # WHEN
    split_excel_into_dirs(sample_excel_file, output_dir, "Dimen", x_filename, "sheet5")

    # Verify files are created for each unique value in "Dimen"
    assert os_path_exists(a_file_path)
    assert os_path_exists(b_file_path)
    assert os_path_exists(c_file_path)

    # Verify contents of one of the created files
    a_df = pandas_read_excel(a_file_path)
    expected_a = DataFrame({"ID": [1, 3], "Dimen": ["A", "A"], "Value": [100, 150]})
    pandas_testing_assert_frame_equal(a_df, expected_a)

    b_df = pandas_read_excel(b_file_path)
    b_expected = DataFrame({"ID": [2, 5], "Dimen": ["B", "B"], "Value": [200, 250]})
    pandas_testing_assert_frame_equal(b_df, b_expected)

    c_df = pandas_read_excel(c_file_path)
    c_expected = DataFrame({"ID": [4], "Dimen": ["C"], "Value": [300]})
    pandas_testing_assert_frame_equal(c_df, c_expected)


def test_split_excel_into_dirs_DoesNotChangeIfColumnIsEmpty(tmp_path, output_dir):
    """Test handling of an empty column."""
    # ESTABLISH Create an Excel file with an empty column
    data = {
        "ID": [1, 2, 3],
        "Dimen": [None, None, None],
        "Value": [100, 200, 300],
    }
    df = DataFrame(data)
    file_path = tmp_path / "empty_column.xlsx"
    df.to_excel(file_path, index=False, sheet_name="sheet5")

    # WHEN
    x_filename = "fizz"
    split_excel_into_dirs(file_path, output_dir, "Dimen", x_filename, "sheet5")

    # THEN Verify that no files are created
    created_files = list(output_dir.iterdir())
    print(f"{created_files=}")
    assert not created_files


def test_split_excel_into_dirs_DoesCreateDirectoryIfColumnEmpty(
    sample_excel_file, tmp_path
):
    """Test if the output directory is created automatically."""
    # ESTABLISH
    output_dir = tmp_path / "nonexistent_output"
    x_filename = "fizz"

    # WHEN
    split_excel_into_dirs(sample_excel_file, output_dir, "Dimen", x_filename, "sheet5")

    # THEN
    assert output_dir.exists()
    print(f"{list(output_dir.iterdir())=}")
    assert list(output_dir.iterdir())


def test_split_excel_into_dirs_SavesToCorrectFileNames(tmp_path, output_dir):
    """Test handling of invalid characters in unique values for filenames."""
    # ESTABLISH Create a DataFrame with special characters in the splitting column
    data = {
        "ID": [1, 2],
        "Dimen": ["A/B", "C\\D"],
        "Value": [100, 200],
    }
    df = DataFrame(data)
    file_path = tmp_path / "special_chars.xlsx"
    df.to_excel(file_path, index=False, sheet_name="sheet5")
    x_filename = "fizz"
    ab_dir = create_path(output_dir, "A_B")
    cd_dir = create_path(output_dir, "C_D")
    b_file_path = create_path(ab_dir, f"{x_filename}.xlsx")
    c_file_path = create_path(cd_dir, f"{x_filename}.xlsx")
    assert os_path_exists(b_file_path) is False
    assert os_path_exists(c_file_path) is False

    # WHEN
    split_excel_into_dirs(file_path, output_dir, "Dimen", x_filename, "sheet5")

    # THEN
    assert os_path_exists(b_file_path)
    assert os_path_exists(c_file_path)


def test_if_nan_return_None_ReturnsObj(env_dir_setup_cleanup):
    # ESTABLISH
    ex1_df = DataFrame([["yao", None]], columns=["face_name", "example_col"])
    ex1_sheet_name = "ex1"
    ex1_filename = "ex1.xlsx"
    ex1_path = create_path(idea_fisc_mstr_dir(), ex1_filename)
    upsert_sheet(ex1_path, ex1_sheet_name, ex1_df)
    gen_df = pandas_read_excel(ex1_path, sheet_name=ex1_sheet_name)
    nan_example = gen_df["example_col"][0]
    print(f"{nan_example=}")
    print(f"{type(nan_example)=}")

    # WHEN / THEN
    assert if_nan_return_None(None) is None
    assert if_nan_return_None("example") == "example"
    assert if_nan_return_None(33) == 33
    assert if_nan_return_None(nan_example) is None

    # ALSO
    print(f"{type(numpy_nan)=}")
    print(f"{type(nan_example)=}")
    print(f"{type(float64(None))=}")
    print(f"{float64(None)=}")
    print(f"  {nan_example=}")
    # assert float64(None) == nan_example
    # assert numpy_nan == nan_example


def test_set_dataframe_first_two_columns_Scenario0_BasicFunctionality():
    # ESTABLISH a DataFrame with three columns
    df = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6], "C": [7, 8, 9]})

    # When we set the first and second columns to specific values
    updated_df = set_dataframe_first_two_columns(df, 100, 200)

    # Then the first and second columns should reflect the new values
    assert (updated_df["A"] == 100).all(), "First column values are incorrect."
    assert (updated_df["B"] == 200).all(), "Second column values are incorrect."


def test_set_dataframe_first_two_columns_Scenario1_TwoColumns():
    # ESTABLISH a DataFrame with exactly two columns
    df = DataFrame({"A": [0, 0], "B": [0, 0]})

    # When we set the first and second columns to specific values
    updated_df = set_dataframe_first_two_columns(df, -1, -2)

    # Then the first and second columns should reflect the new values
    assert (
        updated_df["A"] == -1
    ).all(), "First column values are incorrect for two columns."
    assert (
        updated_df["B"] == -2
    ).all(), "Second column values are incorrect for two columns."


def test_set_dataframe_first_two_columns_Scenario2_MoreThanTwoColumns():
    # ESTABLISH a DataFrame with more than two columns
    df = DataFrame({"A": [1], "B": [2], "C": [3], "D": [4]})

    # When we set the first and second columns to specific values
    updated_df = set_dataframe_first_two_columns(df, 999, 888)

    # Then the first and second columns should reflect the new values
    assert (
        updated_df["A"] == 999
    ).all(), "First column values are incorrect for multiple columns."
    assert (
        updated_df["B"] == 888
    ).all(), "Second column values are incorrect for multiple columns."


def test_set_dataframe_first_two_columns_Scenario3_EmptyDataframe():
    # ESTABLISH an empty DataFrame with two columns
    df = DataFrame({"A": [], "B": []})

    # When we set the first and second columns to specific values
    updated_df = set_dataframe_first_two_columns(df, 42, 43)

    # Then the DataFrame should remain empty
    assert updated_df.empty, "Empty DataFrame should remain empty."


def test_set_dataframe_first_two_columns_Scenario4_LessThanTwoColumns():
    # ESTABLISH a DataFrame with less than two columns
    df = DataFrame({"A": [1, 2, 3]})

    # When we attempt to set the first and second columns
    # Then a ValueError should be raised
    with pytest_raises(ValueError, match="DataFrame must have at least two columns."):
        set_dataframe_first_two_columns(df, 10, 20)


def test_check_dataframe_column_names_ScenarioCorrectColumnNames():
    # ESTABLISH a DataFrame with the correct first two column names
    df = DataFrame({"A": [1, 2, 3], "B": [4, 5, 6]})

    # When we check the first two column names
    result = check_dataframe_column_names(df, "A", "B")

    # Then the result should be True
    assert result == True, "Expected True when column names match."


def test_check_dataframe_column_names_ScenarioIncorrectColumnNames():
    # ESTABLISH a DataFrame with incorrect first two column names
    df = DataFrame({"X": [1, 2, 3], "Y": [4, 5, 6]})

    # When we check the first two column names
    result = check_dataframe_column_names(df, "A", "B")

    # Then the result should be False
    assert result == False, "Expected False when column names do not match."


def test_check_dataframe_column_names_ScenarioPartialColumnMatch():
    # ESTABLISH a DataFrame where only the first column matches
    df = DataFrame({"A": [1, 2, 3], "Y": [4, 5, 6]})

    # When we check the first two column names
    result = check_dataframe_column_names(df, "A", "B")

    # Then the result should be False
    assert result == False, "Expected False when only one column name matches."


def test_check_dataframe_column_names_ScenarioLessThanTwoColumns():
    # ESTABLISH a DataFrame with less than two columns
    df = DataFrame({"A": [1, 2, 3]})

    # When we check the first two column names
    # Then a ValueError should be raised
    with pytest_raises(ValueError, match="DataFrame must have at least two columns."):
        check_dataframe_column_names(df, "A", "B")


def test_update_all_face_name_event_int_columns_Scenario0_UpdatesValidSheet(
    env_dir_setup_cleanup,
):
    # sourcery skip: no-loop-in-tests
    # ESTABLISH
    excel_path = create_path(idea_fisc_mstr_dir(), "test_excel.xlsx")
    set_dir(idea_fisc_mstr_dir())
    yao_str = "Yao"
    event3 = 3
    # A workbook with valid and invalid sheets
    workbook = openpyxl_Workbook()
    ws1 = workbook.active
    validsheet_str = "ValidSheet"
    invalidsheet_str = "InvalidSheet"
    ws1.title = validsheet_str
    ws1.append(["event_int", "face_name", "other"])
    for _ in range(5):
        ws1.append([event3, yao_str, "value4"])

    ws2 = workbook.create_sheet(invalidsheet_str)
    ws2.append(["wrong", "headers", "data"])
    for _ in range(5):
        ws2.append([event3, yao_str, "value3"])

    workbook.save(excel_path)
    bob_str = "Bob"
    event7 = 7
    workbook = openpyxl_load_workbook(excel_path)
    ws1 = workbook[validsheet_str]
    for row in range(2, ws1.max_row + 1):
        assert ws1.cell(row=row, column=1).value != event7
        assert ws1.cell(row=row, column=2).value != bob_str
    ws2 = workbook[invalidsheet_str]
    for row in range(2, ws2.max_row + 1):
        assert ws2.cell(row=row, column=1).value == event3
        assert ws2.cell(row=row, column=2).value == yao_str

    # WHEN: We update the workbook
    update_all_face_name_event_int_columns(excel_path, bob_str, event7)

    # THEN: Only the valid sheet should be updated
    workbook = openpyxl_load_workbook(excel_path)
    ws1 = workbook[validsheet_str]
    for row in range(2, ws1.max_row + 1):
        assert ws1.cell(row=row, column=1).value == event7
        assert ws1.cell(row=row, column=2).value == bob_str
    ws2 = workbook["InvalidSheet"]
    for row in range(2, ws2.max_row + 1):
        assert ws2.cell(row=row, column=1).value == event3
        assert ws2.cell(row=row, column=2).value == yao_str


def test_update_all_face_name_event_int_columns_Scenario1_NoMatchingSheets(
    env_dir_setup_cleanup,
):
    # ESTABLISH: A workbook with no matching headers
    excel_path = create_path(idea_fisc_mstr_dir(), "test_excel.xlsx")
    set_dir(idea_fisc_mstr_dir())
    workbook = openpyxl_Workbook()
    ws = workbook.active
    ws.append(["foo", "bar"])
    workbook.save(excel_path)
    # Ensure initial data is correct
    workbook = openpyxl_load_workbook(excel_path)
    ws = workbook.active
    assert ws.cell(row=1, column=1).value == "foo"
    assert ws.cell(row=1, column=2).value == "bar"

    # WHEN: We attempt to update the workbook
    update_all_face_name_event_int_columns(excel_path, "Bob", 7)

    # THEN: No updates should be made
    workbook = openpyxl_load_workbook(excel_path)
    ws = workbook.active
    assert ws.cell(row=1, column=1).value == "foo"
    assert ws.cell(row=1, column=2).value == "bar"
